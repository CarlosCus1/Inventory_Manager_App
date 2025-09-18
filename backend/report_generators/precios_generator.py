from .base_generator import BaseReportGenerator, autosize_columns, DEFAULT_STYLES
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Any, Dict, List, Optional

class PreciosReportGenerator(BaseReportGenerator):
    def __init__(self, writer: Any, form_data: Dict[str, Any], list_data: List[Dict[str, Any]], data: Optional[Dict[str, Any]] = None, usuario_data: Optional[Dict[str, Any]] = None):
        super().__init__(writer, form_data, list_data, data, usuario_data)
        self.report_type = "COMPARATIVO DE PRECIOS"
        self.report_key = "precios"

    def get_filename(self) -> str:
        """Genera nombre de archivo normalizado"""
        client_name = self.cliente.lower().replace(" ", "_").replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u")
        date_str = datetime.now().strftime("%d-%m-%y")
        return f"comparativo_precios_{client_name}_{date_str}.xlsx"

    def _get_normalized_headers(self) -> List[str]:
        """Encabezados normalizados para el comparativo de precios"""
        return [
            "CODIGO", "EAN13", "EAN14", "NOMBRE PRODUCTO", "BASE (M1)",
            "M2", "DIF. M2", "% M2",
            "M3", "DIF. M3", "% M3",
            "M4", "DIF. M4", "% M4",
            "M5", "DIF. M5", "% M5",
            "PRECIO MIN", "PRECIO MAX", "% MIN", "% MAX",
            "PRECIO SUG.", "DIF. SUG.", "% SUG.",
            "PROMEDIO", "DESVIACIÓN ESTÁNDAR", "DISPERSIÓN", "RANKING DE PRECIO",
            "% DIF. VS PROMEDIO", "% DIF. VS MÍNIMO", "% DIF. VS MÁXIMO",
            "% DIF. VS SUGERIDO", "COMPETIDORES MÁS BARATOS", "COMPETIDORES MÁS CAROS"
        ]

    def generate(self):
        worksheet = self.workbook.create_sheet(title="COMPARATIVO_PRECIOS", index=0)
        if len(self.workbook.sheetnames) > 1 and "Sheet" in self.workbook.sheetnames:
            self.workbook.remove(self.workbook["Sheet"])

        # 1. Datos Generales normalizados
        marcas = [self.form_data.get(f'marca{i}', f'Marca {i}') for i in range(1, 6)]
        doc_type = self.form_data.get('documentType', '').upper()
        doc_num = self.form_data.get('documento_cliente', '')
        doc_display = f"{doc_type}: {doc_num}" if doc_type and doc_num else doc_num

        general_data = {
            "Cliente": self.cliente,
            "Documento": doc_display,
            "Código de Cliente": self.form_data.get('codigo_cliente', ''),
            "Sucursal": self.form_data.get('sucursal') or 'principal',
            "Responsable": self.usuario,
            "Fecha": datetime.now(),
            "Total Productos": len(self.list_data),
            "Marca 1 (Base)": marcas[0],
            "Marca 2": marcas[1],
            "Marca 3": marcas[2],
            "Marca 4": marcas[3],
            "Marca 5": marcas[4],
        }
        table_start_row = self._create_general_data_block(worksheet, general_data, start_row=1) + 1

        # 2. Encabezados de la tabla normalizados
        headers = self._get_normalized_headers()
        
        # Estilos específicos para este reporte
        brand_fills = {
            'M2': PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid'),
            'M3': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
            'M4': PatternFill(start_color='FFDDC1', end_color='FFDDC1', fill_type='solid'),
            'M5': PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid'),
        }
        sugerido_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        minmax_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        kpi_fill = PatternFill(start_color='E0FFFF', end_color='E0FFFF', fill_type='solid') # Light Cyan for new KPIs
        sugerido_font = Font(name='Arial', italic=True)
        
        # Aplicar estilos a los encabezados
        current_header_col = 1
        static_headers_part1 = ["CODIGO", "EAN13", "EAN14", "NOMBRE PRODUCTO"]
        for header_text in static_headers_part1:
            cell = worksheet.cell(row=table_start_row, column=current_header_col, value=header_text)
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            current_header_col += 1

        # BASE (M1) header
        cell = worksheet.cell(row=table_start_row, column=current_header_col, value=f'{marcas[0]} (BASE)')
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        current_header_col += 1

        # M2, M3, M4, M5 headers
        for i in range(1, 5):
            marca_name = marcas[i]
            cell = worksheet.cell(row=table_start_row, column=current_header_col, value=marca_name)
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            cell.fill = brand_fills[f'M{i+1}']
            current_header_col += 1

            cell = worksheet.cell(row=table_start_row, column=current_header_col, value=f'DIF. {marca_name}')
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            cell.fill = brand_fills[f'M{i+1}']
            current_header_col += 1

            cell = worksheet.cell(row=table_start_row, column=current_header_col, value=f'% {marca_name}')
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            cell.fill = brand_fills[f'M{i+1}']
            current_header_col += 1

        # Min/Max headers
        static_headers_minmax = ["PRECIO MIN", "PRECIO MAX", "% MIN", "% MAX"]
        for header_text in static_headers_minmax:
            cell = worksheet.cell(row=table_start_row, column=current_header_col, value=header_text)
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            cell.fill = minmax_fill
            current_header_col += 1

        # Suggested Price headers
        static_headers_sugerido = ["PRECIO SUG.", "DIF. SUG.", "% SUG."]
        for header_text in static_headers_sugerido:
            cell = worksheet.cell(row=table_start_row, column=current_header_col, value=header_text)
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            cell.fill = sugerido_fill
            cell.font = sugerido_font
            current_header_col += 1

        # New KPI headers - Optimizados para espacio horizontal
        static_headers_kpi = [
            "PROMEDIO", "DESV. STD", "DISPERSIÓN", "RANKING",
            "% VS PROM", "% VS MÍN", "% VS MÁX",
            "% VS SUG", "+ BARATOS", "+ CAROS"
        ]
        for header_text in static_headers_kpi:
            cell = worksheet.cell(row=table_start_row, column=current_header_col, value=header_text)
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            cell.fill = kpi_fill
            current_header_col += 1

        # 3. Cuerpo y Fórmulas normalizados
        currency_format = '#,##0.00'
        percentage_format = '0.00%'
        
        current_row = table_start_row + 1
        for item in self.list_data:
            worksheet.cell(row=current_row, column=1, value=self._normalize_value(item.get("codigo")))
            worksheet.cell(row=current_row, column=2, value=self._normalize_value(item.get("cod_ean")))
            worksheet.cell(row=current_row, column=3, value=self._normalize_value(item.get("ean_14")))
            worksheet.cell(row=current_row, column=4, value=self._normalize_value(item.get("nombre")))

            precios_map = item.get("precios", {})
            base_price = precios_map.get(marcas[0])
            base_cell = worksheet.cell(row=current_row, column=5, value=base_price)
            base_cell.number_format = currency_format
            base_cell.alignment = self._get_right_alignment()
            base_cell_coord = base_cell.coordinate

            price_cells_coords = [base_cell_coord]
            col_idx = 6
            for i in range(1, 5):
                marca_name = marcas[i]
                price = precios_map.get(marca_name)
                
                price_cell = worksheet.cell(row=current_row, column=col_idx, value=price)
                price_cell.number_format = currency_format
                price_cell.alignment = self._get_right_alignment()
                price_cell.fill = brand_fills[f'M{i+1}']
                price_cells_coords.append(price_cell.coordinate)

                dif_cell = worksheet.cell(row=current_row, column=col_idx + 1)
                dif_cell.value = f'=IF(ISNUMBER({price_cell.coordinate}), {price_cell.coordinate}-{base_cell_coord}, "")'
                dif_cell.number_format = currency_format
                dif_cell.alignment = self._get_right_alignment()
                dif_cell.fill = brand_fills[f'M{i+1}']

                pct_cell = worksheet.cell(row=current_row, column=col_idx + 2)
                pct_cell.value = f'=IF(AND(ISNUMBER({price_cell.coordinate}), {price_cell.coordinate}<>0), ({base_cell_coord}-{price_cell.coordinate})/{price_cell.coordinate}, 0)'
                pct_cell.number_format = percentage_format
                pct_cell.alignment = self._get_right_alignment()
                pct_cell.fill = brand_fills[f'M{i+1}']
                col_idx += 3

            # --- Helper Columns for Aggregate Functions ---
            helper_col_start = 150  # Start far away to avoid conflicts
            helper_range_coords = []
            for i, price_coord in enumerate(price_cells_coords):
                helper_cell = worksheet.cell(row=current_row, column=helper_col_start + i)
                helper_cell.value = f'=IF(ISNUMBER({price_coord}), {price_coord}, "")'
                helper_range_coords.append(helper_cell.coordinate)
            
            helper_range_str = ""
            if helper_range_coords:
                helper_range_str = f'{helper_range_coords[0]}:{helper_range_coords[-1]}'


            # Min y Max
            min_price_cell = worksheet.cell(row=current_row, column=col_idx)
            min_price_cell.value = f'=IFERROR(MIN(({helper_range_str})), "")'
            min_price_cell.number_format = currency_format
            min_price_cell.alignment = self._get_right_alignment()
            min_price_cell.fill = minmax_fill

            max_price_cell = worksheet.cell(row=current_row, column=col_idx + 1)
            max_price_cell.value = f'=IFERROR(MAX(({helper_range_str})), "")'
            max_price_cell.number_format = currency_format
            max_price_cell.alignment = self._get_right_alignment()
            max_price_cell.fill = minmax_fill

            min_pct_cell = worksheet.cell(row=current_row, column=col_idx + 2)
            min_pct_cell.value = f'=IF(AND(ISNUMBER({base_cell_coord}), {base_cell_coord}<>0), ({min_price_cell.coordinate}/{base_cell_coord})-1, 0)'
            min_pct_cell.number_format = percentage_format
            min_pct_cell.alignment = self._get_right_alignment()
            min_pct_cell.fill = minmax_fill

            max_pct_cell = worksheet.cell(row=current_row, column=col_idx + 3)
            max_pct_cell.value = f'=IF(AND(ISNUMBER({base_cell_coord}), {base_cell_coord}<>0), ({max_price_cell.coordinate}/{base_cell_coord})-1, 0)'
            max_pct_cell.number_format = percentage_format
            max_pct_cell.alignment = self._get_right_alignment()
            max_pct_cell.fill = minmax_fill

            col_idx += 4

            # Precio Sugerido (Manual con fallback a promedio)
            sugerido_manual = item.get('precio_sugerido')
            sugerido_cell = worksheet.cell(row=current_row, column=col_idx)
            
            if isinstance(sugerido_manual, (int, float)):
                sugerido_cell.value = sugerido_manual
            else:
                sugerido_cell.value = f'=IFERROR(AVERAGE(({helper_range_str})), "")'

            sugerido_cell.fill = sugerido_fill
            sugerido_cell.font = sugerido_font
            sugerido_cell.number_format = currency_format
            sugerido_cell.alignment = self._get_right_alignment()

            # DIF. SUG.
            dif_sug_cell = worksheet.cell(row=current_row, column=col_idx + 1)
            dif_sug_cell.value = f'=IF(ISNUMBER({sugerido_cell.coordinate}), {sugerido_cell.coordinate}-{base_cell_coord}, "")'
            dif_sug_cell.fill = sugerido_fill
            dif_sug_cell.font = sugerido_font
            dif_sug_cell.number_format = currency_format
            dif_sug_cell.alignment = self._get_right_alignment()

            # % SUG.
            pct_sug_cell = worksheet.cell(row=current_row, column=col_idx + 2)
            pct_sug_cell.value = f'=IF(AND(ISNUMBER({base_cell_coord}), {base_cell_coord}<>0), ({sugerido_cell.coordinate}/{base_cell_coord})-1, 0)'
            pct_sug_cell.fill = sugerido_fill
            pct_sug_cell.font = sugerido_font
            pct_sug_cell.number_format = percentage_format
            pct_sug_cell.alignment = self._get_right_alignment()
            
            # --- Nuevos KPIs ---
            col_idx += 3 # Mover el índice de columna después de % SUG.

            # PROMEDIO
            avg_cell = worksheet.cell(row=current_row, column=col_idx)
            avg_cell.value = f'=IFERROR(AVERAGE(({helper_range_str})), "")'
            avg_cell.number_format = currency_format
            avg_cell.alignment = self._get_right_alignment()
            avg_cell.fill = kpi_fill
            col_idx += 1

            # DESVIACIÓN ESTÁNDAR
            stdev_cell = worksheet.cell(row=current_row, column=col_idx)
            stdev_cell.value = f'=IFERROR(STDEV({helper_range_str}), "")'
            stdev_cell.number_format = currency_format
            stdev_cell.alignment = self._get_right_alignment()
            stdev_cell.fill = kpi_fill
            col_idx += 1

            # DISPERSIÓN
            dispersion_cell = worksheet.cell(row=current_row, column=col_idx)
            cv_formula = f"IF({avg_cell.coordinate}<>0, {stdev_cell.coordinate}/{avg_cell.coordinate}, 0)"
            dispersion_text_formula = f'IF({cv_formula}>=0.3, "ALTA", IF({cv_formula}>=0.15, "MEDIA", "BAJA"))'
            full_dispersion_formula = f'=IF(ISNUMBER({stdev_cell.coordinate}), {dispersion_text_formula} & " (" & TEXT({cv_formula}, "0.0%") & ")", "")'
            dispersion_cell.value = full_dispersion_formula
            dispersion_cell.alignment = self._get_right_alignment()
            dispersion_cell.fill = kpi_fill
            col_idx += 1

            # RANKING DE PRECIO
            rank_cell = worksheet.cell(row=current_row, column=col_idx)
            rank_formula = f'=IF(ISNUMBER({base_cell_coord}), COUNTIF({helper_range_str}, "<" & {base_cell_coord}) + 1 & "/" & COUNT(({helper_range_str})), "")'
            rank_cell.value = rank_formula
            rank_cell.alignment = self._get_right_alignment()
            rank_cell.fill = kpi_fill
            col_idx += 1

            # % DIF. VS PROMEDIO
            pct_vs_avg_cell = worksheet.cell(row=current_row, column=col_idx)
            pct_vs_avg_cell.value = f'=IF(AND(ISNUMBER({base_cell_coord}), ISNUMBER({avg_cell.coordinate}), {avg_cell.coordinate}<>0), ({base_cell_coord}/{avg_cell.coordinate})-1, 0)'
            pct_vs_avg_cell.number_format = percentage_format
            pct_vs_avg_cell.alignment = self._get_right_alignment()
            pct_vs_avg_cell.fill = kpi_fill
            col_idx += 1

            # % DIF. VS MÍNIMO (nueva columna, diferente a % MIN)
            pct_vs_min_cell = worksheet.cell(row=current_row, column=col_idx)
            pct_vs_min_cell.value = f'=IF(AND(ISNUMBER({base_cell_coord}), ISNUMBER({min_price_cell.coordinate}), {min_price_cell.coordinate}<>0), ({base_cell_coord}/{min_price_cell.coordinate})-1, 0)'
            pct_vs_min_cell.number_format = percentage_format
            pct_vs_min_cell.alignment = self._get_right_alignment()
            pct_vs_min_cell.fill = kpi_fill
            col_idx += 1

            # % DIF. VS MÁXIMO (nueva columna, diferente a % MAX)
            pct_vs_max_cell = worksheet.cell(row=current_row, column=col_idx)
            pct_vs_max_cell.value = f'=IF(AND(ISNUMBER({base_cell_coord}), ISNUMBER({max_price_cell.coordinate}), {max_price_cell.coordinate}<>0), ({base_cell_coord}/{max_price_cell.coordinate})-1, 0)'
            pct_vs_max_cell.number_format = percentage_format
            pct_vs_max_cell.alignment = self._get_right_alignment()
            pct_vs_max_cell.fill = kpi_fill
            col_idx += 1

            # % DIF. VS SUGERIDO (nueva columna, diferente a % SUG.)
            pct_vs_sug_cell = worksheet.cell(row=current_row, column=col_idx)
            pct_vs_sug_cell.value = f'=IF(AND(ISNUMBER({base_cell_coord}), ISNUMBER({sugerido_cell.coordinate}), {sugerido_cell.coordinate}<>0), ({base_cell_coord}/{sugerido_cell.coordinate})-1, 0)'
            pct_vs_sug_cell.number_format = percentage_format
            pct_vs_sug_cell.alignment = self._get_right_alignment()
            pct_vs_sug_cell.fill = kpi_fill
            col_idx += 1

            # COMPETIDORES MÁS BARATOS
            cheaper_competitors_cell = worksheet.cell(row=current_row, column=col_idx)
            cheaper_competitors_cell.value = f'=IF(ISNUMBER({base_cell_coord}), COUNTIF({helper_range_str}, "<" & {base_cell_coord}), "")'
            cheaper_competitors_cell.alignment = self._get_right_alignment()
            cheaper_competitors_cell.fill = kpi_fill
            col_idx += 1

            # COMPETIDORES MÁS CAROS
            more_expensive_competitors_cell = worksheet.cell(row=current_row, column=col_idx)
            more_expensive_competitors_cell.value = f'=IF(ISNUMBER({base_cell_coord}), COUNTIF({helper_range_str}, ">" & {base_cell_coord}), "")'
            more_expensive_competitors_cell.alignment = self._get_right_alignment()
            more_expensive_competitors_cell.fill = kpi_fill
            col_idx += 1            
            current_row += 1

        # Ocultar columnas auxiliares
        if self.list_data:
            num_helper_cols = len(price_cells_coords)
            for i in range(num_helper_cols):
                col_letter = get_column_letter(helper_col_start + i)
                worksheet.column_dimensions[col_letter].hidden = True

        # Aplicar estilos a la tabla usando el método base
        self._apply_table_styles(worksheet, table_start_row, current_row - 1, len(headers))

        # 4. Ajustar columnas
        autosize_columns(worksheet)

    def _get_right_alignment(self) -> Alignment:
        """Devuelve alineación derecha normalizada"""
        return DEFAULT_STYLES['right_alignment']
