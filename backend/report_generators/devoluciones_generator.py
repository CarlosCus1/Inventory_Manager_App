from .base_generator import BaseReportGenerator, STYLE_CONFIG, autosize_columns
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime
import pandas as pd

class DevolucionesReportGenerator(BaseReportGenerator):
    def generate(self):
        config = {
            'sheet_name': "devoluciones",
            'df_cols': ["codigo", "cod_ean", "nombre", "cantidad", "peso", "observaciones"],
            'info_fields': [
                ("hoja de devolucion", ""),
                ("Cliente", self.form_data.get('cliente', '')),
                ("Motivo", self.form_data.get('motivo', '')),
                ("Fecha", self.form_data.get('fecha', ''))
            ],
            'header_start_row': 7,
            'number_formats': {'D': '0', 'E': '#,##0.00'}
        }

        sheet_name = config['sheet_name']
        
        fecha_raw = str(self.form_data.get('fecha', '')).strip()
        try:
            dt = datetime.fromisoformat(fecha_raw.replace('Z', '+00:00'))
            fecha_ddmmyy = dt.strftime('%d-%m-%y')
        except (ValueError, TypeError):
            fecha_ddmmyy = datetime.now().strftime('%d-%m-%y')

        for i, (label, _) in enumerate(config['info_fields']):
            if label == "Fecha":
                config['info_fields'][i] = ("Fecha", fecha_ddmmyy)

        df = pd.DataFrame(self.list_data)
        for col in config['df_cols']:
            if col not in df.columns:
                df[col] = "" if col in ("cod_ean", "observaciones", "nombre", "linea") else 0
        df = df[config['df_cols']]

        data_start_row = config['header_start_row']
        df.to_excel(self.writer, sheet_name=sheet_name, startrow=data_start_row, index=False, header=False)
        ws = self.writer.sheets[sheet_name]

        style_info = STYLE_CONFIG.get('devoluciones', STYLE_CONFIG['default'])

        current_row = 1
        for label, value in config['info_fields']:
            cell = ws.cell(row=current_row, column=1, value=label)
            self._apply_header_style(cell, style_info)
            if value:
                ws.cell(row=current_row, column=2, value=value)
            current_row += 1

        for idx, h in enumerate(config['df_cols'], start=1):
            cell = ws.cell(row=data_start_row, column=idx, value=h)
            self._apply_header_style(cell, style_info)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        data_first_row_idx = data_start_row + 1
        for r_idx in range(data_first_row_idx, data_first_row_idx + len(df)):
            for col_letter, fmt in config['number_formats'].items():
                ws[f'{col_letter}{r_idx}'].number_format = fmt

        total_row_idx = data_start_row + len(df) + 2
        cell = ws.cell(row=total_row_idx, column=1, value="totales")
        self._apply_header_style(cell, style_info)

        if len(df) > 0:
            start_row_data_formula = data_start_row + 1
            end_row_data_formula = data_start_row + len(df)
            total_cantidad_cell = ws.cell(row=total_row_idx, column=4)
            total_cantidad_cell.value = f"=SUM(D{start_row_data_formula}:D{end_row_data_formula})"
            total_cantidad_cell.number_format = "0"
            total_peso_cell = ws.cell(row=total_row_idx, column=5)
            total_peso_cell.value = f"=SUMPRODUCT(D{start_row_data_formula}:D{end_row_data_formula},E{start_row_data_formula}:E{end_row_data_formula})"
            total_peso_cell.number_format = "#,##0.00"
        else:
            ws.cell(row=total_row_idx, column=4, value=0).number_format = "0"
            ws.cell(row=total_row_idx, column=5, value=0).number_format = "#,##0.00"
        
        autosize_columns(ws)
