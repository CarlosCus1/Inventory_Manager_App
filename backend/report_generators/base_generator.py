from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, NamedStyle
from typing import Any, Dict, List, Optional
from datetime import datetime

STYLE_CONFIG = {
    "devoluciones": {"bg_color": "FFC7CE"},  # Rojo claro
    "pedido": {"bg_color": "ADD8E6"},  # Azul claro
    "inventario": {"bg_color": "90EE90"},  # Verde claro
    "precios": {"bg_color": "FFD966"},  # Amarillo claro
    "default": {"bg_color": "D9D9D9"},
}

# Estilos nombrados para consistencia
DEFAULT_STYLES = {
    'header_font': Font(name='Arial', size=11, bold=True),
    'body_font': Font(name='Arial', size=10),
    'totals_font': Font(name='Arial', size=11, bold=True, color="FFFFFF"),
    'header_fill': PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
    'totals_fill': PatternFill(start_color="404040", end_color="404040", fill_type="solid"),
    'thin_border': Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')),
    'center_alignment': Alignment(horizontal="center", vertical="center"),
    'left_alignment': Alignment(horizontal="left", vertical="center"),
    'right_alignment': Alignment(horizontal="right", vertical="center"),
}

class BaseReportGenerator:
    def __init__(self, writer: Any, form_data: Dict[str, Any], list_data: List[Dict[str, Any]], data: Optional[Dict[str, Any]] = None, usuario_data: Optional[Dict[str, Any]] = None):
        self.writer = writer
        self.form_data = form_data
        self.list_data = list_data
        self.data = data
        self.usuario_data = usuario_data if usuario_data else {}
        self.workbook = writer.book
        self.report_type = "default"
        self.report_key = "default"
        self.cliente = self.form_data.get('cliente', 'N/A')
        self.usuario = self.usuario_data.get('nombre', 'N/A')

    def get_filename(self) -> str:
        """Genera nombre de archivo normalizado"""
        report_name = self.report_type.lower().replace(" ", "_").replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u")
        client_name = self.cliente.lower().replace(" ", "_").replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u")
        date_str = datetime.now().strftime("%d-%m-%y")
        return f"{report_name}_{client_name}_{date_str}.xlsx"

    def _create_general_data_block(self, worksheet: Worksheet, general_data: Dict[str, Any], start_row: int) -> int:
        """Crea el bloque de datos generales con un título y formato mejorado."""
        
        # 1. Título de la sección
        title_cell = worksheet.cell(row=start_row, column=1, value="DATOS GENERALES")
        title_cell.font = Font(name='Arial', size=12, bold=True)
        title_cell.fill = DEFAULT_STYLES['header_fill']
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        worksheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=2)
        
        # Aplicar borde a las celdas fusionadas
        top_left_cell = worksheet.cell(row=start_row, column=1)
        top_right_cell = worksheet.cell(row=start_row, column=2)
        top_left_cell.border = Border(top=Side(style='thin'), left=Side(style='thin'), bottom=Side(style='thin'))
        top_right_cell.border = Border(top=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin'))

        row = start_row + 1

        # 2. Datos clave-valor
        style_info = STYLE_CONFIG.get(self.report_key, STYLE_CONFIG["default"])
        fill = PatternFill(start_color=style_info["bg_color"], end_color=style_info["bg_color"], fill_type="solid")

        for key, value in general_data.items():
            # Key
            key_cell = worksheet.cell(row=row, column=1, value=self._normalize_text(key))
            key_cell.font = DEFAULT_STYLES['header_font']
            key_cell.fill = fill
            key_cell.border = DEFAULT_STYLES['thin_border']
            key_cell.alignment = DEFAULT_STYLES['left_alignment']
            
            # Value
            value_cell = worksheet.cell(row=row, column=2, value=self._normalize_value(value))
            value_cell.font = DEFAULT_STYLES['body_font']
            value_cell.border = DEFAULT_STYLES['thin_border']
            value_cell.alignment = DEFAULT_STYLES['left_alignment']
            row += 1
        
        # Fila en blanco como separador
        return row + 1

    def _normalize_text(self, text: Any) -> str:
        """Normaliza texto para consistencia"""
        if text is None:
            return ""
        return str(text).strip()

    def _normalize_value(self, value: Any) -> Any:
        """Normaliza valores según su tipo"""
        if value is None:
            return ""
        elif isinstance(value, bool):
            return "Sí" if value else "No"
        elif isinstance(value, (int, float)):
            return value
        elif isinstance(value, datetime):
            return value.strftime("%d/%m/%Y")
        else:
            return self._normalize_text(value)

    def _apply_table_styles(self, worksheet: Worksheet, start_row: int, end_row: int, end_col: int):
        """Aplica estilos normalizados a toda la tabla"""
        # Estilo de encabezados
        for col in range(1, end_col + 1):
            cell = worksheet.cell(row=start_row, column=col)
            cell.font = DEFAULT_STYLES['header_font']
            cell.fill = DEFAULT_STYLES['header_fill']
            cell.border = DEFAULT_STYLES['thin_border']
            cell.alignment = DEFAULT_STYLES['center_alignment']

        # Estilo del cuerpo
        for row in worksheet.iter_rows(min_row=start_row + 1, max_row=end_row, min_col=1, max_col=end_col):
            for cell in row:
                cell.font = DEFAULT_STYLES['body_font']
                cell.border = DEFAULT_STYLES['thin_border']
                if isinstance(cell.value, (int, float)):
                    cell.alignment = DEFAULT_STYLES['right_alignment']
                else:
                    cell.alignment = DEFAULT_STYLES['left_alignment']

    def _apply_totals_style(self, row: int, start_col: int, end_col: int, worksheet: Worksheet):
        """Aplica estilo normalizado a la fila de totales"""
        for col in range(start_col, end_col + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.font = DEFAULT_STYLES['totals_font']
            cell.fill = DEFAULT_STYLES['totals_fill']
            cell.border = DEFAULT_STYLES['thin_border']
            cell.alignment = DEFAULT_STYLES['center_alignment']

    def _get_normalized_headers(self) -> List[str]:
        """Método para ser sobrescrito por cada generador específico"""
        return []

    def validate_and_process(self, item: Dict[str, Any]) -> List[str]:
        errors = []
        if item.get('cantidad', 0) <= 0:
            errors.append("Cantidad inválida")
        if not item.get('codigo'):
            errors.append("Código requerido")
        return errors

    def generate(self):
        raise NotImplementedError("Cada generador debe implementar su propio método 'generate'.")

def autosize_columns(worksheet: Worksheet):
    """Ajusta el ancho de las columnas de forma mejorada, ignorando celdas fusionadas."""
    from openpyxl.cell import MergedCell

    for col in worksheet.columns:
        max_length = 0
        
        column_letter = None
        for cell in col:
            if not isinstance(cell, MergedCell):
                column_letter = cell.column_letter
                break
        
        if not column_letter:
            continue

        has_numbers = False
        
        for cell in col:
            if isinstance(cell, MergedCell):
                continue

            try:
                if cell.value:
                    if isinstance(cell.value, (int, float)):
                        has_numbers = True
                    
                    font_size = cell.font.sz if cell.font and cell.font.sz else 11
                    is_bold = cell.font.b if cell.font and cell.font.b else False
                    
                    if is_bold:
                        multiplier = 1.3
                    elif has_numbers:
                        multiplier = 1.1
                    else:
                        multiplier = 1.0
                    
                    cell_len = len(str(cell.value)) * multiplier
                    if cell_len > max_length:
                        max_length = cell_len
            except:
                pass
        
        min_width = 12
        adjusted_width = max(min_width, (max_length + 2))
        worksheet.column_dimensions[column_letter].width = min(adjusted_width, 30)
