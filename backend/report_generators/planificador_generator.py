from .base_generator import BaseReportGenerator, STYLE_CONFIG, autosize_columns
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime
import pandas as pd
import unicodedata

class PlanificadorReportGenerator(BaseReportGenerator):
    def generate(self):
        data = self.data # Use the 'data' attribute passed to the constructor
        chart_color_name = self.form_data.get('linea_planificador_color') # Get chart color from form_data

        COLOR_MAP = STYLE_CONFIG.get("planner_charts", {})
        hex_color = COLOR_MAP.get(chart_color_name or "default", COLOR_MAP.get("default", "4472C4"))

        def _lighten_color(hex_color, factor=0.5):
            hex_color = hex_color.lstrip('#')
            rgb = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
            new_rgb = [int(val + (255 - val) * factor) for val in rgb]
            return f'{new_rgb[0]:02X}{new_rgb[1]:02X}{new_rgb[2]:02X}'

        def format_month_year_es(date_obj: datetime) -> str:
            MONTH_NAMES_ES = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
            return f"{MONTH_NAMES_ES[date_obj.month - 1]} {date_obj.year}"

        info_data = [
            ("Cód. Cliente:", data.get('codigoCliente', '')),
            ("RUC:", data.get('ruc', '')),
            ("Cliente:", data.get('razonSocial', '')),
            ("Línea:", data.get('linea', '')),
            ("Cód. Pedido:", data.get('pedido', '')),
            ("Monto Total:", data.get('montoOriginal', 0)),
            ("Total Letras:", len(data.get('fechasOrdenadas', [])))
        ]
        df_info = pd.DataFrame(info_data, columns=["Campo", "Valor"])

        resumen_mensual = data.get('resumenMensual', {})
        df_resumen = pd.DataFrame(list(resumen_mensual.items()), columns=["Mes", "Monto (S/)"])
        df_resumen["Porcentaje"] = df_resumen["Monto (S/)"] / data.get('montoOriginal', 1)

        fechas_ordenadas = data.get('fechasOrdenadas', [])
        montos_asignados = data.get('montosAsignados', {})
        detalle_data = [{
            "N°": i + 1,
            "Fecha de Vencimiento": fecha,
            "Monto (S/)": montos_asignados.get(fecha, 0)
        } for i, fecha in enumerate(fechas_ordenadas)]
        df_detalle = pd.DataFrame(detalle_data)

        df_info.to_excel(self.writer, sheet_name="Reporte Dashboard", startrow=2, index=False, header=False)
        df_resumen.to_excel(self.writer, sheet_name="Reporte Dashboard", startrow=len(df_info) + 4, index=False)
        df_detalle.to_excel(self.writer, sheet_name="Detalle de Pagos", index=False)

        ws_dashboard = self.writer.sheets["Reporte Dashboard"]
        ws_detalle = self.writer.sheets["Detalle de Pagos"]

        color_hex_fill = STYLE_CONFIG['pedido']['header_color'] # Using 'pedido' color for main fill as in original
        
        # Main title style
        ws_dashboard.merge_cells('A1:C1')
        cell = ws_dashboard['A1']
        cell.value = "DISTRIBUCION DE MONTOS POR FECHA"
        cell.font = Font(bold=True, size=16, color="FFFFFF")
        cell.fill = PatternFill(start_color=color_hex_fill, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Detalle sheet column width
        for col in ws_detalle.columns:
            ws_detalle.column_dimensions[col[0].column_letter].width = 20

        # Chart creation and styling
        chart = BarChart()
        chart.title = "Resumen de Montos por Mes"
        chart.style = 12
        chart.y_axis.title = 'Monto (S/)'
        chart.x_axis.title = 'Mes'
        chart.legend = None

        data_ref = Reference(ws_dashboard, min_col=2, min_row=len(df_info) + 5, max_row=len(df_info) + 5 + len(df_resumen))
        cats_ref = Reference(ws_dashboard, min_col=1, min_row=len(df_info) + 6, max_row=len(df_info) + 5 + len(df_resumen))

        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        series = chart.series[0]
        fill = PatternFill(patternType='solid', fgColor=hex_color)
        series.graphicalProperties = GraphicalProperties(solidFill=fill)

        ws_dashboard.add_chart(chart, "E3")

        cliente = str(data.get('razonSocial', '')).strip()
        fecha_raw = str(data.get('fechasOrdenadas', [''])[0]).strip()
        try:
            dt = datetime.fromisoformat(fecha_raw.replace('Z', '+00:00'))
            fecha_ddmmyy = dt.strftime('%d-%m-%y')
        except (ValueError, TypeError):
            fecha_ddmmyy = datetime.now().strftime('%d-%m-%y')

        safe_cliente = unicodedata.normalize('NFKD', cliente).encode('ascii', 'ignore').decode('utf-8').strip().replace(" ", "_") or "sin_cliente"
        self.filename = f"planificador_{safe_cliente}_{fecha_ddmmyy}.xlsx"
